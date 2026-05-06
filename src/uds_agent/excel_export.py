"""JSON 测试结果 → Excel 文件导出。

每个服务生成一个 Sheet，格式匹配 Het_UDS_TestSpecification 模板。
"""

from __future__ import annotations

import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# 列宽配置（匹配模板）
COL_WIDTHS = {
    "A": 12,  # Sequence Number
    "B": 18,  # System Requirement ID
    "C": 18,  # Case ID
    "D": 55,  # Case Name
    "E": 8,   # Priority
    "F": 10,  # Author
    "G": 16,  # Design Method
    "H": 14,  # Precondition
    "I": 50,  # Test Procedure
    "J": 50,  # Expected Output
    "K": 12,  # Actual Output
    "L": 10,  # (empty)
    "M": 12,  # Defect ID
    "N": 10,
    "O": 10,
    "P": 10,
}

# 样式
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SECTION_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def export_to_excel(services_data: list[dict]) -> bytes:
    """将多个服务的测试用例导出为 Excel 字节流。

    Args:
        services_data: API 返回的 services 数组

    Returns:
        .xlsx 文件的字节内容
    """
    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    for svc in services_data:
        sheet_name = _safe_sheet_name(svc.get("sheet_name", "Sheet1"))
        ws = wb.create_sheet(title=sheet_name)

        # 设置列宽
        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        # 写表头
        _write_header(ws)

        # 写数据
        test_cases = svc.get("test_cases", [])
        row = 4  # 数据从第4行开始
        current_section = ""
        current_subsection = ""
        seq = 0

        for tc in test_cases:
            section = tc.get("section", "")
            subsection = tc.get("subsection", "")

            # 写 section header
            if section != current_section:
                current_section = section
                current_subsection = ""
                _write_section_row(ws, row, section)
                row += 1

            # 写 subsection header
            if subsection != current_subsection:
                current_subsection = subsection
                _write_section_row(ws, row, subsection)
                row += 1

            # 无测试用例的子功能只显示标题
            if tc.get("is_empty_section"):
                continue

            # 写 test case row
            seq += 1
            _write_test_case_row(ws, row, seq, tc)
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_header(ws):
    """写3行表头，匹配 Het_UDS_TestSpecification 模板。"""
    _style = lambda c: (
        setattr(c, 'font', HEADER_FONT) or
        setattr(c, 'fill', HEADER_FILL) or
        setattr(c, 'border', THIN_BORDER)
    )

    # Row 1
    ws["B1"] = "Test Level"
    _style(ws["B1"])
    ws["C1"] = "SYS.5"
    _style(ws["C1"])

    # Row 2 & 3: 合并单元格
    ws.merge_cells("A2:A3")
    ws["A2"] = "Sequence Number"
    _style(ws["A2"]); _style(ws["A3"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B2:B3")
    ws["B2"] = "System Requirement ID"
    _style(ws["B2"]); _style(ws["B3"])
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("C2:F2")
    ws["C2"] = "Test Case"
    _style(ws["C2"]); _style(ws["F2"])
    ws["C2"].alignment = Alignment(horizontal="center", vertical="center")

    # C3-F3 子标题
    for col, h in [(3, "ID"), (4, "Name"), (5, "Priority"), (6, "Author")]:
        cell = ws.cell(row=3, column=col, value=h)
        _style(cell)
        cell.alignment = Alignment(horizontal="center")

    # G-M: 各列第2、3行合并
    merge_headers = {
        7: "Design Method", 8: "Precondition", 9: "Test Procedure",
        10: "Expected Output", 11: "Actual Output", 12: "Judgement Result", 13: "Defect ID",
    }
    for col, title in merge_headers.items():
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
        cell = ws.cell(row=2, column=col, value=title)
        _style(cell)
        _style(ws.cell(row=3, column=col))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_section_row(ws, row, text):
    """写 section/subsection 标题行。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER


def _write_test_case_row(ws, row, seq, tc):
    """写一条测试用例。"""
    values = [
        seq,                                    # A: Sequence Number
        tc.get("system_requirement_id", ""),    # B: System Requirement ID
        tc.get("case_id", ""),                  # C: Case ID
        tc.get("case_name", ""),                # D: Case Name
        tc.get("priority", "High"),             # E: Priority
        tc.get("author", ""),                   # F: Author
        tc.get("design_method", ""),            # G: Design Method
        tc.get("precondition", ""),             # H: Precondition
        tc.get("test_procedure", ""),           # I: Test Procedure
        tc.get("expected_output", ""),          # J: Expected Output
        tc.get("actual_output", ""),            # K: Actual Output
        "",                                     # L
        tc.get("defect_id", ""),                # M: Defect ID
        "", "", "",                             # N-P
    ]

    for col, val in enumerate(values, 1):
        if isinstance(val, str):
            val = val.replace("<br>", "\n").replace("<br/>", "\n").replace("<BR>", "\n")
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if col == 1:
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.alignment = WRAP_ALIGN


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """确保 Sheet 名合法（最长31字符，不含特殊字符）。"""
    for ch in ("\\", "/", "*", "?", ":", "[", "]"):
        name = name.replace(ch, "_")
    return name[:max_len]
